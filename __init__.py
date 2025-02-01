import logging
import os
import json
import requests
import csv
import tempfile
from openpyxl import load_workbook
import azure.functions as func
from azure.storage.blob import BlobServiceClient

# Constants from application settings
BING_MAPS_KEY = os.environ["BING_MAPS_KEY"]
EXCEL_CONTAINER = os.environ["EXCEL_CONTAINER"]
EXCEL_FILENAME = os.environ["EXCEL_FILENAME"]

# Set up Blob Storage connection (the connection string is in the FUNCTIONS_EXTENSION_VERSION settings)
BLOB_CONN_STR = os.environ["AzureWebJobsStorage"]

# Bing Maps Local Search API URL template
# We will use the Bing Maps REST API Search endpoint:
# https://dev.virtualearth.net/REST/v1/LocalSearch/?query={query}&key={BING_MAPS_KEY}
BING_SEARCH_URL = "https://dev.virtualearth.net/REST/v1/LocalSearch/"

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('LookupNonprofits function triggered.')

    try:
        # Connect to Blob Storage
        blob_service_client = BlobServiceClient.from_connection_string(BLOB_CONN_STR)
        container_client = blob_service_client.get_container_client(EXCEL_CONTAINER)

        # Download the Excel file to a temporary file
        temp_excel_path = os.path.join(tempfile.gettempdir(), EXCEL_FILENAME)
        with open(temp_excel_path, "wb") as download_file:
            download_stream = container_client.download_blob(EXCEL_FILENAME)
            download_file.write(download_stream.readall())

        # Open the Excel file and load the first 200 rows
        wb = load_workbook(filename=temp_excel_path, data_only=True)
        ws = wb.active

        # Prepare a list to hold result rows: each row will have Organisation, Phone, Website
        results = []
        header = ["Organisation", "Phone", "Website"]
        results.append(header)

        # Assume that the first row is header; process next 200 rows
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row_count >= 200:
                break
            org_name = row[0]  # Organisation column
            city = row[2]      # Ort column (city)
            if org_name is None:
                continue

            # Formulate a search query: organization name and city
            query = f"{org_name} {city}"
            params = {
                "query": query,
                "key": BING_MAPS_KEY
            }
            # Call the Bing Maps Local Search API
            response = requests.get(BING_SEARCH_URL, params=params)
            if response.status_code != 200:
                logging.error(f"Error from Bing Maps API for query {query}: {response.text}")
                phone = ""
                website = ""
            else:
                data = response.json()
                # Navigate the response JSON structure.
                # Look for a resource in the first result that has 'PhoneNumber' and 'Website'
                phone = ""
                website = ""
                try:
                    resources = data["resourceSets"][0]["resources"]
                    if resources:
                        # Pick the first result that appears to match the query
                        result = resources[0]
                        phone = result.get("PhoneNumber", "")
                        website = result.get("Website", "")
                    else:
                        logging.info(f"No resources found for query: {query}")
                except Exception as e:
                    logging.error(f"Error parsing Bing response for query {query}: {e}")

            results.append([org_name, phone, website])
            row_count += 1

        # Write results to a CSV file
        output_filename = "NonprofitLookupResults.csv"
        temp_csv_path = os.path.join(tempfile.gettempdir(), output_filename)
        with open(temp_csv_path, mode="w", newline="", encoding="utf-8") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerows(results)

        # Upload the CSV file back to the same container
        blob_client = container_client.get_blob_client(output_filename)
        with open(temp_csv_path, "rb") as data_file:
            blob_client.upload_blob(data_file, overwrite=True)

        return func.HttpResponse(
            json.dumps({"message": "Lookup completed successfully", "outputFile": output_filename}),
            mimetype="application/json",
            status_code=200
        )
    except Exception as e:
        logging.error(f"Error in function execution: {e}")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            mimetype="application/json",
            status_code=500
        )
